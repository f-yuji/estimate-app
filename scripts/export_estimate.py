from __future__ import annotations
import argparse,json,shutil,subprocess
from pathlib import Path
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
TEMPLATE=ROOT/"templates"/"estimate-template-fixed.xlsm"
MAPPING=json.loads((ROOT/"config"/"excel-cell-mapping.json").read_text(encoding="utf-8"))
def export(data_path:Path,output:Path,pdf:bool=False):
    if not TEMPLATE.exists(): raise FileNotFoundError("Excelテンプレートが見つかりません。")
    data=json.loads(data_path.read_text(encoding="utf-8"));shutil.copy2(TEMPLATE,output)
    wb=load_workbook(output,keep_vba=True)
    for key in ("estimate.customer_name","estimate.site_address","estimate.issue_date","estimate.total_amount"):
        cfg=MAPPING[key];wb[cfg["sheet"]][cfg["cell"]]=data.get(key)
    cfg=MAPPING["estimate_items"];capacity=cfg["end_row"]-cfg["start_row"]+1
    if len(data.get("estimate_items",[]))>capacity*len(cfg["sheets"]):raise ValueError("Excelテンプレートの明細行数を超えています。")
    for i,item in enumerate(data.get("estimate_items",[])):
        ws=wb[cfg["sheets"][i//capacity]];row=cfg["start_row"]+i%capacity
        for field,col in cfg["columns"].items():ws[f"{col}{row}"]=item.get(field)
    wb.calculation.fullCalcOnLoad=True;wb.calculation.forceFullCalc=True;wb.save(output)
    if pdf:
        ps=(ROOT/"scripts"/"export_pdf.ps1").resolve();subprocess.run(["powershell","-ExecutionPolicy","Bypass","-File",str(ps),"-InputPath",str(output)],check=True)
if __name__=="__main__":
    p=argparse.ArgumentParser();p.add_argument("data",type=Path);p.add_argument("output",type=Path);p.add_argument("--pdf",action="store_true");a=p.parse_args();export(a.data,a.output,a.pdf)
