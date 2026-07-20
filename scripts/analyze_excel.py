from __future__ import annotations

import json
import zipfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / "templates" / "estimate-template.xlsm"
OUT = ROOT / "analysis"


def main() -> None:
    OUT.mkdir(exist_ok=True)
    wb = load_workbook(BOOK, data_only=False, keep_vba=True)
    cached = load_workbook(BOOK, data_only=True, keep_vba=True)
    report: dict[str, object] = {"file": str(BOOK), "sheets": [], "errors": []}
    formulas: list[dict[str, object]] = []

    for ws in wb.worksheets:
        sheet = {
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
            "print_area": str(ws.print_area) if ws.print_area else None,
            "print_title_rows": ws.print_title_rows,
            "orientation": ws.page_setup.orientation,
            "paper_size": ws.page_setup.paperSize,
            "fit_to_width": ws.page_setup.fitToWidth,
            "fit_to_height": ws.page_setup.fitToHeight,
            "formula_count": 0,
        }
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raw_formula = cell.value
                    formula = getattr(raw_formula, "text", str(raw_formula))
                    item = {
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "formula": formula,
                        "cached_value": cached[ws.title][cell.coordinate].value,
                    }
                    formulas.append(item)
                    sheet["formula_count"] += 1
                    for marker in ("#REF!", "#VALUE!", "#NAME?"):
                        if marker in formula:
                            report["errors"].append({**item, "marker": marker})
        report["sheets"].append(sheet)

    with zipfile.ZipFile(BOOK) as archive:
        names = archive.namelist()
        report["archive"] = {
            "has_vba_project": "xl/vbaProject.bin" in names,
            "has_custom_ui": any(n.startswith("customUI/") for n in names),
            "has_activex": any("activeX" in n for n in names),
            "has_drawings": any(n.startswith("xl/drawings/") for n in names),
            "has_images": any(n.startswith("xl/media/") for n in names),
            "extensions": dict(Counter(Path(n).suffix for n in names)),
        }

    (OUT / "workbook_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    (OUT / "formulas.json").write_text(
        json.dumps(formulas, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
