from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "templates" / "estimate-template.xlsm"
OUTPUT = ROOT / "templates" / "estimate-template-fixed.xlsm"


def formula_for(column: str, row: int) -> str:
    if column == "K":
        return (
            f'=IFERROR(INDEX(物件概要!$L$3:$L$50,MATCH(1,'
            f'(物件概要!$K$3:$K$50<>"")*ISNUMBER(SEARCH(物件概要!$K$3:$K$50,B{row})),0)),'
            f'IFERROR(INDEX(\'物件概要(2)\'!$L$3:$L$49,MATCH(1,'
            f'(\'物件概要(2)\'!$K$3:$K$49<>"")*ISNUMBER(SEARCH(\'物件概要(2)\'!$K$3:$K$49,B{row})),0)),""))'
        )
    return (
        f'=IFERROR(INDEX(\'物件概要(2)\'!$N$3:$N$49,MATCH(1,'
        f'(\'物件概要(2)\'!$M$3:$M$49<>"")*ISNUMBER(SEARCH(\'物件概要(2)\'!$M$3:$M$49,B{row})),0)),"")'
    )


def main() -> None:
    wb = load_workbook(SOURCE, keep_vba=True, data_only=False)
    for sheet_name in ("内訳明細書", "内訳明細書 (2)"):
        ws = wb[sheet_name]
        for cell_ref in ("K10", "M10", "K11", "M11"):
            cell = ws[cell_ref]
            formula = formula_for(cell.column_letter, cell.row)
            current = cell.value
            if hasattr(current, "text"):
                current.text = formula
            else:
                cell.value = formula
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
